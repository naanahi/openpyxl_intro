"""既存のExcelファイルからデータを読み込み
"""

from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename="output.xlsx")

sheet_name = wb.sheetnames[0]
ws1 = wb[sheet_name]
x = ws1['A4'].value
print(x)
ws2 = wb.create_sheet('newsheet') ## 新しいシートの作成
ws2['A1'] = x
wb.save('output.xlsx')