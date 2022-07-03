"""グラフを作成する
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
from openpyxl.chart import BarChart, Reference

wb = load_workbook(filename="output.xlsx")
sheet_name = wb.sheetnames[0]
ws1 = wb[sheet_name] ## シートオブジェクトの取得

x = Reference(ws1, min_col=2, min_row=2, max_col=2, max_row=13)
data = Reference(ws1, min_col=3, min_row=2, max_col=3, max_row=13)
chart = BarChart()
chart.add_data(data)

chart.set_categories(x)
ws1.add_chart(chart, "E2")
wb.save("output.xlsx")