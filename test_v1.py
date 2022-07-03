"""新規のExcelファイルに書き込み
"""

from openpyxl import Workbook
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws["A4"] = 10
cel = ws["A4"] ## 変更したいセルのオブジェクトを取得
cel.font = Font(size=12, bold=True)
wb.save('output.xlsx') ## ファイルへの出力