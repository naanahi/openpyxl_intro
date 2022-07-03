"""自由に配列操作をしてみる
"""

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border
from openpyxl.chart import BarChart, Reference

wb = load_workbook(filename="matrix.xlsx")
sheet_name = wb.sheetnames[0]
ws1 = wb[sheet_name] ## シートオブジェクトの取得

## 情報の読み込み
value_lists = []
for row in ws1.iter_rows(min_row=2, min_col=1):
    # 空のリストを作成
    value_list = []
    # 1行分のデータをリストに格納する。
    for c in row:
        value_list.append(c.value)
    # リスト内のデータを表示させる
    value_lists.append(value_list)

## 情報の出力
for i in value_lists:
    ws2 = wb.create_sheet(i[0])
    count = 1
    for j in i:
        ws2.cell(row = count, column = 1, value = j)    
        count += 1

wb.save("save.xlsx")