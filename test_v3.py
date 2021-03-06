"""セルの結合～罫線を引く
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Side, Border

from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename="output.xlsx")
sheet_name = wb.sheetnames[0]
ws1 = wb[sheet_name] ## シートオブジェクトの取得

ws1.merge_cells('A5:A7')
ws1['A5'] = "こんにちは"

s = Side(style='thin')
b = Border(left=s, right=s, top=s, bottom=s)
cell = ws1['B2']
cell.border = b

wb.save('output.xlsx')